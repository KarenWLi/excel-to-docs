from __future__ import print_function
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from openpyxl import load_workbook

filename = "spreadsheet.xlsx" ## Put the title of the Excel sheet here
wb = load_workbook(filename, data_only=True)
sh = wb["Sheet1"]

begin = 0
end = len(filename)-5
title = filename[begin:end]

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/documents']

creds = None

if os.path.exists('token.pickle'):
    with open('token.pickle', 'rb') as token:
        creds = pickle.load(token)

if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    else:
        flow = InstalledAppFlow.from_client_secrets_file(
            'credentials.json', SCOPES)
        creds = flow.run_local_server(port=0)
    with open('token.pickle', 'wb') as token:
        pickle.dump(creds, token)

service = build('docs', 'v1', credentials=creds)

body = {
'title': title
}
doc = service.documents() \
    .create(body=body).execute()
docId = doc.get('documentId')
print = docId

row = 2 

requests = [
        {
            'insertText': {
                'endOfSegmentLocation': {
                    'segmentId': ''
                },
                'text': title + '\n \n'
            },
        },
        
    ]

while sh.cell(row,1).value is not None:
    col = 1
    text = ""
    request = {}


    while col < 7:
        text = text + sh.cell(1,col).value + ": " + str(sh.cell(row,col).value) + "\n"
        col = col + 1

    
    request = {
        'insertText': {
            'endOfSegmentLocation': {
                'segmentId': ''
            },
            'text': text + "\n \n \n \n \n" 
        },
    }

    requests.append(request)

  
    row = row +1

document = service.documents().batchUpdate(documentId=docId, body={'requests': requests}).execute()

#Improvable aspects: Currently unable to make a heading or page break with Docs API, sad times